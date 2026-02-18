
import logging
import json
import base64
import os
import requests
import pandas as pd
import threading
from urllib.parse import urlparse, parse_qs
from typing import Dict, Any, Optional, List
from groq import Groq, RateLimitError
from tenacity import retry, stop_after_attempt, wait_exponential, retry_if_exception_type
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PIL import Image, ImageEnhance, ImageFilter, ImageOps
import io
import time

# Import Prompts
from src.prompts import (
    UNIVERSAL_PARSER_PROMPT,
    SERIAL_EXTRACTION_TEMPLATE,
    SOLAR_MODULE_RULES,
    GENERIC_ASSET_RULES,
    COMMON_SERIAL_RULES,
    HUMAN_DETECTION_PROMPT
)

# Configure Enterprise Logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger("SolarAI.Engine")

class SolarVisionEngine:
    """
    Enterprise-Grade Computer Vision Engine utilizing Groq's LPUs.
    Features:
    - Dynamic Document Parsing (Aadhaar, PAN, Solar Labels)
    - Rate Limiting with Exponential Backoff
    - Robust JSON Structure Enforcement
    - Two-Pass Serial Extraction Logic (Raw -> Forensic)
    """
    
    def __init__(self, api_key: Optional[str] = None, model_id: str = "llama-3.2-90b-vision-preview"):
        """Initialize the Groq client securely from Env or Argument."""
        self.api_key = api_key or os.getenv("GROQ_API_KEY")
        if not self.api_key:
            raise ValueError("GROQ_API_KEY is not set. Please provide it or set the environment variable.")
            
        self.client = Groq(api_key=self.api_key)
        # Using variable model ID for flexibility, default fallback from env if provided there
        self.model_id = os.getenv("GROQ_MODEL_ID", model_id)

    def _encode_image(self, image_path: str) -> str:
        """Optimized Base64 encoding with aggressive resizing for speed."""
        with Image.open(image_path) as img:
            # Resize if too large (e.g. > 1280px) to reduce payload size
            # This makes upload to Groq 3x faster
            if img.width > 1280 or img.height > 1280:
                img.thumbnail((1280, 1280))
            
            buffered = io.BytesIO()
            # Convert to RGB to avoid mode issues with some formats
            if img.mode != 'RGB':
                img = img.convert('RGB')
                
            img.save(buffered, format="JPEG", quality=85)
            return base64.b64encode(buffered.getvalue()).decode('utf-8')

    def _preprocess_forensic(self, image_path: str) -> str:
        """
        Forensic Image Enhancement for Serial Number Extraction.
        Applies: Grayscale -> High Contrast -> Sharpening
        """
        try:
            with Image.open(image_path) as img:
                # 0. Handle EXIF Rotation (Crucial for mobile uploads)
                img = ImageOps.exif_transpose(img)
                
                # 1. Convert to Grayscale to remove color noise
                img = img.convert('L')
                
                # 2. Enhance Contrast (Make text blacker against background)
                enhancer = ImageEnhance.Contrast(img)
                img = enhancer.enhance(2.0)  # Double the contrast
                
                # 3. Enhance Sharpness (Define edges of digits)
                enhancer = ImageEnhance.Sharpness(img)
                img = enhancer.enhance(2.0)  # Significant sharpening
                
                # 4. Resize if too small (width < 1000px) for better OCR
                if img.width < 1000:
                    ratio = 1000 / img.width
                    new_size = (1000, int(img.height * ratio))
                    img = img.resize(new_size, Image.Resampling.LANCZOS)

                # Convert to base64
                buffered = io.BytesIO()
                img.save(buffered, format="JPEG", quality=95)
                return base64.b64encode(buffered.getvalue()).decode('utf-8')
        except Exception as e:
            logger.warning(f"Preprocessing failed: {e}. Falling back to raw image.")
            return self._encode_image(image_path)

    @retry(
        retry=retry_if_exception_type(RateLimitError),
        stop=stop_after_attempt(5),
        wait=wait_exponential(multiplier=1, min=2, max=20)
    )
    def analyze_installation(self, image_path: str) -> Dict[str, Any]:
        """
        Executes a high-performance vision analysis with automatic rate-limit handling.
        Dynamically extracts key-value pairs based on document type.
        """
        logger.info(f"Initiating Vision Analysis on: {image_path}")
        try:
            base64_image = self._encode_image(image_path)
            
            completion = self.client.chat.completions.create(
                model=self.model_id,
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {"type": "text", "text": UNIVERSAL_PARSER_PROMPT},
                            {
                                "type": "image_url",
                                "image_url": {
                                    "url": f"data:image/jpeg;base64,{base64_image}"
                                }
                            }
                        ]
                    }
                ],
                temperature=0.1,           # Low temperature for deterministic output
                max_tokens=1024,
                top_p=1,
                stream=False               # DISABLED STREAMING for lower latency
            )

            # Get full response immediately
            full_response = completion.choices[0].message.content
                
            logger.debug(f"Raw Model Output: {full_response[:100]}...")

            # Robust JSON Parsing with Fallback Logic
            try:
                # Sanitization: Locate the first '{' and last '}' to handle potential preamble leaks
                json_start = full_response.find('{')
                json_end = full_response.rfind('}') + 1
                if json_start != -1 and json_end != -1:
                    clean_json = full_response[json_start:json_end]
                    return json.loads(clean_json)
                else:
                    raise ValueError("No JSON structure found in response")
            except json.JSONDecodeError:
                logger.error("JSON Parsing Failed. Attempting simplified extraction.")
                return {"error": "Invalid Metadata Format", "raw": full_response}

        except Exception as e:
            logger.error(f"Critical Vision Engine Failure: {str(e)}", exc_info=True)
            return {"error": str(e)}

    @retry(
        retry=retry_if_exception_type(RateLimitError),
        stop=stop_after_attempt(5),
        wait=wait_exponential(multiplier=1, min=2, max=20)
    )
    def extract_serial_only(self, image_path: str, asset_type: str = "Solar Module") -> str:
        """
        Specialized extraction for Serial Numbers with STRICT Validation.
        Uses a 2-Pass Logic (Raw -> Forensic) to Balance Speed and Accuracy.
        """
        def call_model(img_b64: str, rigorous: bool = False) -> str:
            intensity = "EXTREME" if rigorous else "HIGH"
            
            # Construct Prompt based on Asset Type
            prompt = SERIAL_EXTRACTION_TEMPLATE.format(asset_type=asset_type, intensity=intensity)
            
            if asset_type == "Solar Module":
                prompt += SOLAR_MODULE_RULES
            else:
                prompt += GENERIC_ASSET_RULES
                
            prompt += COMMON_SERIAL_RULES
            
            completion = self.client.chat.completions.create(
                model=self.model_id,
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {"type": "text", "text": prompt},
                            {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{img_b64}"}}
                        ]
                    }
                ],
                temperature=0.0,
                max_tokens=60,
                stream=False
            )
            return completion.choices[0].message.content.strip()

        try:
            # Pass 1: RAW IMAGE (Best for modern VLMs like Llama 3.2)
            # We explicitly AVOID pre-processing here to let the model see the natural image.
            # Convert raw file to base64
            base64_v1 = self._encode_image(image_path)
            result_v1 = call_model(base64_v1, rigorous=False)
            
            clean_v1 = ''.join(filter(str.isalnum, result_v1))
            
            # Validation Logic
            if asset_type == "Solar Module":
                # STRICT 16-digit rule for Modules (Immediate Success)
                if len(clean_v1) == 16:
                    return clean_v1
            else:
                # Flexible rule for Inverters/Others
                if len(clean_v1) >= 10: 
                    return clean_v1
                
            # Pass 2: If Raw Image failed strict validation, try 'Forensic Enhancement'
            # This is our fallback for bad lighting/low contrast.
            logger.info(f"Pass 1 (Raw) result '{clean_v1}' suspicious. Retrying Pass 2 (Forensic)...")
            
            # Use the forensic pre-processor (Grayscale + Contrast + Sharpness)
            base64_v2 = self._preprocess_forensic(image_path)
            result_v2 = call_model(base64_v2, rigorous=True) # Send base64 directly
            clean_v2 = ''.join(filter(str.isalnum, result_v2))
            
            # Final Decision
            if asset_type == "Solar Module":
                if len(clean_v2) == 16:
                    return clean_v2
                
                # If neither passed strict check, use heuristics
                diff_v1 = abs(len(clean_v1) - 16)
                diff_v2 = abs(len(clean_v2) - 16)
                
                # Return closest to 16
                if diff_v1 <= diff_v2 and clean_v1:
                    return clean_v1
                return clean_v2 if clean_v2 else (clean_v1 if clean_v1 else "N/A")
            else:
                # Standard logic (Longest plausible string)
                if len(clean_v2) > len(clean_v1):
                    return clean_v2
                return clean_v1 if clean_v1 else "N/A"

        except Exception as e:
            logger.error(f"Serial Extraction Failed: {str(e)}")
            return "Error"

    @retry(
        retry=retry_if_exception_type(RateLimitError),
        stop=stop_after_attempt(5),
        wait=wait_exponential(multiplier=1, min=2, max=20)
    )
    def detect_human(self, image_path: str) -> bool:
        """
        Detects if a human (person/face) is present in the image.
        Returns True/False.
        """
        try:
            base64_image = self._encode_image(image_path)
            
            completion = self.client.chat.completions.create(
                model=self.model_id,
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {"type": "text", "text": HUMAN_DETECTION_PROMPT},
                            {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{base64_image}"}}
                        ]
                    }
                ],
                temperature=0.0,
                max_tokens=10,
                stream=False
            )
            ans = completion.choices[0].message.content.strip().upper()
            return "YES" in ans
        except Exception as e:
            logger.error(f"Human Detection Failed: {e}")
            return False

class BatchProcessor:
    """
    Handles bulk processing of Excel sheets containing Google Drive links.
    Features:
    - Thread-safe Write Operations (concurrent.futures friendly)
    - Connection Pooling (requests.Session)
    - Integrates with SolarVisionEngine for row-by-row analysis.
    """
    def __init__(self, excel_path: str, api_key: Optional[str] = None, model_id: str = "llama-3.2-90b-vision-preview"):
        self.excel_path = excel_path
        # Initialize Engine (it handles env variable fallback)
        self.engine = SolarVisionEngine(api_key=api_key, model_id=model_id)
        
        # Thread Safety Lock for concurrent writes
        self.write_lock = threading.Lock()
        
        # Connection Pool Session for faster downloads
        self.session = requests.Session()
        
        # Load Excel with Pandas for processing logic
        try:
            self.df = pd.read_excel(excel_path)
            # FIX: Cast all to object to prevent "incompatible dtype" warnings when writing strings to empty float columns
            self.df = self.df.astype(object)
            
            # Load Excel with OpenPyXL for style preservation
            self.wb = load_workbook(excel_path)
            self.ws = self.wb.active
            
            # Map column names to 1-based indices for OpenPyXL
            # Assumption: Headers are in the first row
            self.col_map_xl = {}
            for idx, col_cell in enumerate(self.ws[1], start=1):
                if col_cell.value:
                    self.col_map_xl[col_cell.value] = idx
                    
            logger.info(f"Loaded Excel file: {excel_path} with {len(self.df)} rows.")
        except Exception as e:
            logger.error(f"Failed to load Excel file: {e}")
            raise

    def save(self, output_path: str):
        """Saves the processed workbook, preserving styles but auto-fitting columns."""
        try:
            # Auto-resize columns
            for i, col in enumerate(self.ws.columns, start=1):
                col_letter = get_column_letter(i)
                max_length = 0
                
                for cell in col:
                    try:
                        if cell.value:
                            val_len = len(str(cell.value))
                            if val_len > max_length:
                                max_length = val_len
                    except:
                        pass
                
                # Add padding and cap width
                adjusted_width = (max_length + 2)
                # Cap at 60 chars for readability
                if adjusted_width > 60: 
                    adjusted_width = 60
                
                self.ws.column_dimensions[col_letter].width = adjusted_width

            self.wb.save(output_path)
            logger.info(f"Saved processed file to {output_path}")
        except Exception as e:
            logger.error(f"Failed to save output file: {e}")
            raise

    def _update_xl(self, row_idx, col_name, value):
        """Updates the OpenPyXL worksheet cell, preserving style. Thread-safe."""
        try:
            # Pandas index is 0-based. Excel rows are 1-based.
            # Usually row 1 is header. So data starts at row 2.
            # Row in Excel = row_idx + 2
            xl_row = row_idx + 2
            
            if col_name in self.col_map_xl:
                xl_col = self.col_map_xl[col_name]
                with self.write_lock:
                    self.ws.cell(row=xl_row, column=xl_col).value = value
        except Exception as e:
            logger.warning(f"Failed to update Excel cell: {e}")

    def _get_drive_id(self, url: str) -> Optional[str]:
        """Extracts Google Drive file ID from URL."""
        if not isinstance(url, str):
            return None
            
        parsed = urlparse(url)
        if 'drive.google.com' in parsed.netloc:
            # Handle /file/d/ID/view format
            path_parts = parsed.path.split('/')
            if 'd' in path_parts:
                try:
                    return path_parts[path_parts.index('d') + 1]
                except IndexError:
                    pass
            
            # Handle ?id=ID format
            query = parse_qs(parsed.query)
            if 'id' in query:
                return query['id'][0]
                
        return None

    def _download_image(self, drive_id: str, save_path: str) -> bool:
        """Downloads image from Google Drive using ID and Connection Pooling."""
        url = f"https://drive.google.com/uc?export=download&id={drive_id}"
        try:
            # Use self.session for pooled connection
            response = self.session.get(url, stream=True, timeout=10)
            if response.status_code == 200:
                with open(save_path, 'wb') as f:
                    for chunk in response.iter_content(1024):
                        f.write(chunk)
                
                # Validate that we actually downloaded an image (not an HTML error page)
                try:
                    with Image.open(save_path) as img:
                        img.verify() # Simple check for corruption
                    return True
                except Exception:
                    logger.warning(f"Downloaded file {drive_id} is not a valid image.")
                    return False
            else:
                logger.warning(f"Failed to download image: {drive_id} (Status: {response.status_code})")
                return False
        except Exception as e:
            logger.error(f"Error downloading image {drive_id}: {e}")
            return False

    def process_solar_audit_row(self, index, row, status_callback=None):
        """
        Specialized processor for Solar Audit Sheets.
        1. Extract Module Serial Numbers.
        2. Extract Inverter Serial Numbers.
        3. Detect Humans in Installation Photos.
        """
        cols = row.index.tolist()
        count = 0
        audit_remarks = []
        
        # --- Helper 1: Process Serial Number Pairs ---
        def process_pair(photo_col, target_col, asset_type):
            img_url = row.get(photo_col)
            if not img_url or pd.isna(img_url) or not isinstance(img_url, str):
                return False
                
            drive_id = self._get_drive_id(img_url)
            if not drive_id:
                val = "Invalid Link"
                with self.write_lock:
                    self.df.at[index, target_col] = val
                self._update_xl(index, target_col, val)
                audit_remarks.append(f"{target_col}: Invalid Link")
                return False
                
            test_path = f"temp_{asset_type.replace(' ','_')}_{index}.jpg"  
            if self._download_image(drive_id, test_path):
                serial = self.engine.extract_serial_only(test_path, asset_type)
                with self.write_lock:
                    self.df.at[index, target_col] = serial
                self._update_xl(index, target_col, serial)
                
                if serial == "Error":
                    audit_remarks.append(f"{target_col}: Extraction Error")
                
                if os.path.exists(test_path): os.remove(test_path)
                return True
            else:
                 val = "Download Failed"
                 with self.write_lock:
                     self.df.at[index, target_col] = val
                 self._update_xl(index, target_col, val)
                 audit_remarks.append(f"{target_col}: Download Failed")
                 return False

        # --- Helper 2: Check Human Presence ---
        human_detected_overall = False

        def check_human_in_col(col_name, label):
            nonlocal human_detected_overall
            img_url = row.get(col_name)
            if not img_url or pd.isna(img_url):
                audit_remarks.append(f"{label}: Not Present")
                return

            drive_id = self._get_drive_id(img_url)
            if not drive_id:
                audit_remarks.append(f"{label}: Invalid Link")
                return

            temp_path = f"temp_human_{index}_{label}.jpg"
            if self._download_image(drive_id, temp_path):
                has_human = self.engine.detect_human(temp_path)
                if has_human:
                    human_detected_overall = True
                    audit_remarks.append(f"{label}: Human Found")
                else:
                    audit_remarks.append(f"{label}: Clear")
                
                if os.path.exists(temp_path): os.remove(temp_path)
            else:
                audit_remarks.append(f"{label}: Download Failed")

        # 1. Iterate columns for Serial Numbers
        for col in cols:
            # A. Module Serial Numbers
            if "Module Serial Number Photo" in col:
                suffix = col.replace("Module Serial Number Photo", "").strip()
                target_candidates = [c for c in cols if f"AI Module Serial Number {suffix}" in c]
                if target_candidates:
                    if process_pair(col, target_candidates[0], "Solar Module"):
                        count += 1
            
            # B. Inverter Serial Numbers
            elif "Inverter Serial Number Photo" in col:
                suffix = col.replace("Inverter Serial Number Photo", "").strip()
                target_candidates = [c for c in cols if f"AI Inverter Serial Number{suffix}" in c]
                if not target_candidates:
                     target_candidates = [c for c in cols if "AI Inverter Serial Number" in c]
                
                if target_candidates:
                    if process_pair(col, target_candidates[0], "Inverter"):
                        count += 1

        # 2. Check Installation Photos for Humans
        panel_view_cols = [c for c in cols if "Installation Photo" in c and "Panel View" in c]
        inverter_view_cols = [c for c in cols if "Installation Photo" in c and "Inverter View" in c]

        if panel_view_cols: check_human_in_col(panel_view_cols[0], "Panel View")
        if inverter_view_cols: check_human_in_col(inverter_view_cols[0], "Inverter View")

        # 3. Update Human Detection Columns
        human_col_candidates = [c for c in cols if "Human Detected" in c]
        if human_col_candidates:
            val = "YES" if human_detected_overall else "NO"
            with self.write_lock:
                self.df.at[index, human_col_candidates[0]] = val
            self._update_xl(index, human_col_candidates[0], val)

        # 4. Update Remarks
        remarks_col_candidates = [c for c in cols if "ai remarks" in c.lower()]
        if remarks_col_candidates and audit_remarks:
            val = " | ".join(audit_remarks)
            with self.write_lock:
                self.df.at[index, remarks_col_candidates[0]] = val
            self._update_xl(index, remarks_col_candidates[0], val)

        if status_callback:
            if count > 0:
                status_callback(f"✅ Row {index+1}: Extracted {count} Serials | Humans: {human_detected_overall}")
            else:
                status_callback(f"⚠️ Row {index+1}: Processed (No Serials Found)")

    def process_row(self, index, row, img_col, human_col, ai_remarks_col, data_col, status_callback=None):
        """Processes a single row for generic documents."""
        # ... (Old Logic - Unchanged) ...
        image_url = row.get(img_col)
        
        if not image_url or pd.isna(image_url):
            logger.info(f"Row {index}: No image URL found. Skipping.")
            return

        drive_id = self._get_drive_id(image_url)
        if not drive_id:
            logger.warning(f"Row {index}: Could not extract Drive ID from URL: {image_url}")
            return
            
        temp_img_path = f"temp_batch_{index}.jpg"
        
        if self._download_image(drive_id, temp_img_path):
            try:
                # Analyze image
                result = self.engine.analyze_installation(temp_img_path)
                
                # Check for errors
                if "error" in result:
                    val = f"Error: {result['error']}"
                    with self.write_lock:
                        self.df.at[index, ai_remarks_col] = val
                    self._update_xl(index, ai_remarks_col, val)
                    
                    if status_callback: status_callback(f"⚠️ Row {index+1}: Vision Error")
                else:
                    # Fill columns
                    human_detected_val = "YES" if result.get('human_detected', False) else "NO"
                    with self.write_lock:
                        self.df.at[index, human_col] = human_detected_val
                    self._update_xl(index, human_col, human_detected_val)
                    
                    # AI Remarks Logic
                    remarks = []
                    if not result.get('domain_relevant', True):
                        remarks.append(f"Irrelevant: {result.get('rejection_reason', 'Unknown')}")
                    
                    remarks.append(f"Type: {result.get('document_type', 'Unknown')}")
                    
                    # Human Count
                    h_count = result.get('human_count', 0)
                    if h_count > 0:
                         remarks.append(f"Humans: {h_count}")

                    remarks_val = " | ".join(remarks)
                    with self.write_lock:
                        self.df.at[index, ai_remarks_col] = remarks_val
                    self._update_xl(index, ai_remarks_col, remarks_val)
                    
                    # Extracted Data (JSON Dump or Summary)
                    excluded_keys = ['document_type', 'domain_relevant', 'rejection_reason', 'human_detected', 'human_count', 'confidence', 'error']
                    data_only = {k: v for k, v in result.items() if k not in excluded_keys}
                    data_val = str(data_only)
                    with self.write_lock:
                        self.df.at[index, data_col] = data_val
                    self._update_xl(index, data_col, data_val) 
                    
                    if status_callback: status_callback(f"✅ Row {index+1}: Processed")
                
                logger.info(f"Row {index}: Processed successfully.")
                
            except Exception as e:
                logger.error(f"Row {index}: Processing error: {e}")
                val = f"Processing Error: {str(e)}"
                with self.write_lock:
                    self.df.at[index, ai_remarks_col] = val
                self._update_xl(index, ai_remarks_col, val)
                if status_callback: status_callback(f"❌ Row {index+1}: Exception")
            finally:
                if os.path.exists(temp_img_path):
                    os.remove(temp_img_path)
        else:
            val = "Download Failed"
            with self.write_lock:
                self.df.at[index, ai_remarks_col] = val
            self._update_xl(index, ai_remarks_col, val)
            if status_callback: status_callback(f"⚠️ Row {index+1}: Download Failed (Check Permissions)")
